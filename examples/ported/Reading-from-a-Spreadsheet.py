"""Port of AlibreScript ``Reading-from-a-Spreadsheet.py``.

Original: https://help.alibre.com/articles/#!alibre-help-v28/reading-from-a-spreadsheet

Pure-Python ``openpyxl`` usage, no AlibreX involvement. Run with::

    pip install openpyxl
    python Reading-from-a-Spreadsheet.py path\\to\\Book1.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-not-found]


_DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent / "_sample_files" / "sample.xlsx"
)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path, nargs="?", default=_DEFAULT_XLSX,
                        help=f"Path to a .xlsx (default: {_DEFAULT_XLSX.name})")
    args = parser.parse_args()
    if not args.xlsx_path.exists():
        raise SystemExit(f"xlsx not found: {args.xlsx_path}")

    wb = load_workbook(filename=str(args.xlsx_path))
    sheet = wb.active   # first sheet, original used hardcoded "Sheet1"
    assert sheet is not None, f"Workbook {args.xlsx_path.name} has no active sheet"
    print(f"Workbook: {args.xlsx_path.name}  Sheet: {sheet.title}")
    print("Contents:")
    for row in sheet.iter_rows(values_only=True):
        print(" ", row)


if __name__ == "__main__":
    main()
